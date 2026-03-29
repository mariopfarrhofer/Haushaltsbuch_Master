import openpyxl
from datetime import datetime, timedelta

def create_test_excel():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create 2023 sheet
    ws2023 = wb.create_sheet("2023")
    ws2023.append(["Datum", "Zählerstand", "Aktion"])
    
    base_date = datetime(2023, 1, 15)
    val = 1450.0
    
    for i in range(12):
        d_str = base_date.strftime("%d.%m.%Y")
        
        # In summer months simulate "open" for lower unit
        action = ""
        if i == 5:
            action = "open"
        elif i == 8:
            action = "close"
            
        ws2023.append([d_str, round(val, 2), action])
        
        base_date += timedelta(days=30)
        # Normal monthly is ~9.5. When open, add extra ~6-8
        if 5 <= i < 8:
            val += 16.5
        else:
            val += 9.5
    
    # Create 2024 sheet
    ws2024 = wb.create_sheet("2024")
    ws2024.append(["Datum", "Zählerstand", "Aktion"])
    
    base_date = datetime(2024, 1, 15)
    for i in range(12):
        d_str = base_date.strftime("%d.%m.%Y")
        
        action = ""
        if i == 4:
            action = "open"
        elif i == 9:
            action = "close"
            
        ws2024.append([d_str, round(val, 2), action])
        
        base_date += timedelta(days=30)
        if 4 <= i < 9:
            val += 18.2
        else:
            val += 10.2
        
    wb.save("test_data.xlsx")
    print("test_data.xlsx created successfully.")

if __name__ == "__main__":
    create_test_excel()
